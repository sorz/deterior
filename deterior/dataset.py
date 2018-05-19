from collections import namedtuple, defaultdict, OrderedDict
from csv import DictReader
from datetime import datetime
from typing import TextIO, BinaryIO, Dict, Set
from configparser import ConfigParser
import sys
from openpyxl import load_workbook


Record = namedtuple('Record', ['s0', 's1', 't'])
Inspect = namedtuple('Inspect', ['id', 'state', 'date'])


class ExcelReader:
    """Like csv.DictReader, but read MS Excel file.
    """
    def __init__(self, xlsfile: BinaryIO):
        book = load_workbook(xlsfile, read_only=True)
        self.sheet = book.worksheets[0]  # use the first sheet only
        self.fieldnames = [n.value.strip() for n in self.sheet[1]]
        self.line_num = 1
        self._rows = self.sheet.iter_rows(min_row=2)

    def __iter__(self):
        return self

    def __next__(self):
        self.line_num += 1
        row = next(self._rows)
        row = [c.value for c in row]
        return OrderedDict(zip(self.fieldnames, row))


class DataSetReader:
    """Read inspection log file according to the format config.
    """
    def __init__(self, config: TextIO = None) -> None:
        cfg = ConfigParser(interpolation=None)
        cfg.read_dict({
            'ID': {'column': 'ID'},
            'State': {'column': 'State'},
            'Time': {
                'column': 'Time',
                'format': '%Y-%m-%d',
                'unit': '1d',
            },
        })
        if config is not None:
            cfg.read_file(config)
        self.col_id = cfg['ID']['column']
        self.col_state = cfg['State']['column']
        self.col_time = cfg['Time']['column']
        self.time_format = cfg['Time']['format']
        self.time_unit = _time_unit_to_days(cfg['Time']['unit'])

    def _load(self, rows) -> ([Record], int):
        inpsects = []
        for row in rows:
            sid = row.get(self.col_id)
            state = row.get(self.col_state)
            date = row.get(self.col_time)
            if not sid:
                raise ValueError(f'ID not found in row {rows.line_num}')
            if not state or not date:
                print(f'In row {rows.line_num}, {sid} '
                      'has blank state or time, ignored.', file=sys.stderr)
                continue
            if isinstance(date, str):
                date = datetime.strptime(date, self.time_format)
            inpsects.append(Inspect(sid, state, date))
        return self._inpsects_to_records(inpsects)

    def load_csv(self, csvfile: TextIO) -> ([Record], int):
        csv = DictReader(csvfile)
        return self._load(csv)

    def load_xls(self, xlsfile: BinaryIO) -> ([Record], int):
        xls = ExcelReader(xlsfile)
        return self._load(xls)

    def _inpsects_to_records(self, inspects: [Inspect]) -> ([Record], int):
        """Return list of records and the total number of states."""
        states = set()
        id_inspects = defaultdict(list)
        for sid, state, date in inspects:
            states.add(state)
            id_inspects[sid].append((state, date))
        states = _map_states(states)
        print(f'Found {len(states)} states in total')

        records = []
        for states_dates in id_inspects.values():
            if len(states_dates) < 2:
                continue
            states_dates = sorted(states_dates)
            for i in range(len(states_dates) - 1):
                (s0, t0), (s1, t1) = states_dates[i], states_dates[i+1]
                time = round((t1 - t0).days / self.time_unit)
                if time <= 0:
                    continue
                records.append(Record(states[s0], states[s1], time))
        return records, len(states)


def _map_states(states: Set[str]) -> Dict[str, int]:
    """Map a set of string to numerical states."""
    states = sorted(states)
    smap = {s: n for n, s in enumerate(states)}
    for state in states:
        if state != str(smap[state]):
            print(f'Mapping state "{state}" to S{smap[state]}')
    return smap


def _time_unit_to_days(time_unit: str) -> int:
    """Convert "N" to N, "Nd" to N, "Nm" to N * 30, and "Ny" to N * 365.
    """
    if time_unit.isdigit():
        return int(time_unit)
    unit = time_unit[-1].lower()
    num = time_unit[:-1]
    if not num.isdigit() or unit not in 'dmy':
        raise ValueError(f'Time unit "{time_unit}" must be "[0-9]+[dmy]?"')
    num = int(num)
    if unit == 'm':
        num *= 30
    elif unit == 'y':
        num *= 365
    return num
